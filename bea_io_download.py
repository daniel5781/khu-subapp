"""
BEA Input-Output 데이터 일괄 다운로드 스크립트 (v4 — Use/Import/Domestic × Sector/Summary/Detail)
==================================================================================================
미국 BEA(Bureau of Economic Analysis)의 Input-Output 표를
- 3개 표 종류 (USE = Total / IMPORT = 수입행렬 / DOMESTIC = 국산사용표) ×
- 3개 분류 수준 (SECTOR / SUMMARY / DETAIL)
모두 API로 자동 다운로드하고, **종류·수준별로 별도 Excel 워크북**으로 저장합니다.

3종 × 3수준
-----------
USE      = "Use of Commodities by Industries"                            ← 한국은행 A표 총거래표
IMPORT   = "Use of Imported Commodities by Industries"                   ← 한국은행 A표 수입거래표 (Aᵐ)
DOMESTIC = "Use of Domestically Produced Commodities by Industries"      ← 한국은행 A표 국산거래표 (Aᵈ)

DETAIL 수준은 보통 벤치마크 연도(1997, 2002, 2007, 2012, 2017)만 제공됩니다.

산출물
------
- bea_use_table_all_years_sector.xlsx
- bea_use_table_all_years_summary.xlsx
- bea_use_table_all_years_detail.xlsx
- bea_import_matrix_all_years_sector.xlsx
- bea_import_matrix_all_years_summary.xlsx
- bea_import_matrix_all_years_detail.xlsx
- bea_domestic_use_all_years_sector.xlsx
- bea_domestic_use_all_years_summary.xlsx
- bea_domestic_use_all_years_detail.xlsx
- raw/<type>_<level>/bea_YYYY.json    (연도별 원본 JSON 캐시)
- bea_io_download.log

사용법
------
1) pip install requests pandas openpyxl
2) 아래 API_KEY 에 본인의 BEA API 키
3) 필요한 TABLE_TYPES / LEVELS 만 켜고 실행:
       python bea_io_download.py
"""

from __future__ import annotations

import json
import logging
import os
import re
import time
import urllib.request
from pathlib import Path

import pandas as pd
import requests
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# python-dotenv 가 설치돼 있으면 .env 자동 로드. 없어도 환경변수만 직접 export
# 해두면 동작한다.
try:
    from dotenv import load_dotenv
    load_dotenv(Path(__file__).resolve().parent / ".env")
except ImportError:
    pass

# ==================================================================
# 설정
# ==================================================================
API_KEY = os.getenv("BEA_API_KEY", "").strip()
if not API_KEY:
    raise RuntimeError(
        "BEA_API_KEY 환경변수가 설정되지 않았습니다.\n"
        "  · 저장소 루트에 `.env` 파일을 만들고  BEA_API_KEY=<발급받은 키>  를 한 줄 추가하거나,\n"
        "  · 쉘에서  export BEA_API_KEY=<발급받은 키>  로 직접 설정하세요.\n"
        "BEA 키는 https://apps.bea.gov/api/signup/ 에서 무료 발급."
    )
BASE_URL = "https://apps.bea.gov/api/data/"
DATASET = "InputOutput"

START_YEAR = 2000
END_YEAR = None                                 # None → 최신 연도까지

# 받고 싶은 종류·수준만 남기면 됨.
# IMPORT 매트릭스(Aᵐ)는 BEA API 가 노출하지 않으므로 download_bea_static_files() 가
# 별도로 처리한다 (apps.bea.gov/industry/xls/io-annual/ImportMatrices_*.xlsx).
TABLE_TYPES = ["USE", "MAKE", "TR_IXI", "TR_IXC", "TR_CXC"]
LEVELS      = ["SECTOR", "SUMMARY", "DETAIL"]

OUTPUT_DIR = Path(__file__).resolve().parent
RAW_ROOT = OUTPUT_DIR / "raw"
RAW_ROOT.mkdir(exist_ok=True)

REQUEST_TIMEOUT = 60
SLEEP_BETWEEN_CALLS = 1.0   # 분당 100회 rate-limit 보호

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s | %(levelname)s | %(message)s",
    handlers=[
        logging.FileHandler(OUTPUT_DIR / "bea_io_download.log", encoding="utf-8"),
        logging.StreamHandler(),
    ],
)
log = logging.getLogger("bea_io")


# ==================================================================
# 표 식별 규칙
# ==================================================================
# 분류 수준은 desc 안의 다음 토큰을 찾아 매칭
LEVEL_TOKENS = {
    "SECTOR":  ["SECTOR", "SECT", "- SEC"],
    "SUMMARY": ["SUMMARY", " SUM"],
    "DETAIL":  ["DETAIL", " DET"],
}

# 표 종류 매칭 규칙 — include 의 모든 토큰이 desc(uppercase)에 들어있어야 매칭(AND).
# BEA InputOutput API 가 실제로 노출하는 표는 4가지 계열이다 (PDF App. K 기준):
# - USE    : "The Use of Commodities by Industries"        — 총거래표 (수입 포함)
# - MAKE   : "The Domestic Supply of Commodities by Industries" — 공급/Make 표 (commodity × industry)
# - TR_IXI : "Total Requirements, Industry-by-Industry"    — 정사각 (I−A)⁻¹ 산업×산업
# - TR_IXC : "Total Requirements, Industry-by-Commodity"   — 산업×재화
# - TR_CXC : "Total Requirements, Commodity-by-Commodity"  — 재화×재화
TYPE_RULES = {
    "USE":    {"include": ["USE OF COMMODITIES"], "exclude": []},
    "MAKE":   {"include": ["DOMESTIC SUPPLY"], "exclude": []},
    "TR_IXI": {"include": ["TOTAL REQUIREMENTS", "INDUSTRY-BY-INDUSTRY"], "exclude": []},
    "TR_IXC": {"include": ["TOTAL REQUIREMENTS", "INDUSTRY-BY-COMMODITY"], "exclude": []},
    "TR_CXC": {"include": ["TOTAL REQUIREMENTS", "COMMODITY-BY-COMMODITY"], "exclude": []},
}

# 출력 파일명 prefix
TYPE_FILE_PREFIX = {
    "USE":    "bea_use_table",
    "MAKE":   "bea_supply_table",
    "TR_IXI": "bea_total_req_industry_by_industry",
    "TR_IXC": "bea_total_req_industry_by_commodity",
    "TR_CXC": "bea_total_req_commodity_by_commodity",
}

# 워크시트 타이틀에 들어갈 사람 친화적 이름
TYPE_HUMAN_NAME = {
    "USE":    "Use of Commodities by Industries (Total Use, incl. imports)",
    "MAKE":   "Domestic Supply of Commodities by Industries (Make / Supply Table)",
    "TR_IXI": "Total Requirements, Industry-by-Industry (I−A)⁻¹",
    "TR_IXC": "Total Requirements, Industry-by-Commodity",
    "TR_CXC": "Total Requirements, Commodity-by-Commodity",
}


# ==================================================================
# API helper
# ==================================================================
def _get(params: dict) -> dict:
    q = {"UserID": API_KEY, "ResultFormat": "json", **params}
    resp = requests.get(BASE_URL, params=q, timeout=REQUEST_TIMEOUT)
    resp.raise_for_status()
    data = resp.json()
    err = data.get("BEAAPI", {}).get("Error")
    if isinstance(err, dict):
        raise RuntimeError(f"BEA API Error: {err}")
    return data


def _unwrap_results(data: dict) -> dict:
    res = data["BEAAPI"]["Results"]
    if isinstance(res, list):
        if not res:
            raise RuntimeError("빈 Results 리스트")
        res = res[0]
    if not isinstance(res, dict):
        raise RuntimeError(f"예상치 못한 Results 형태: {type(res).__name__}")
    if "Error" in res:
        raise RuntimeError(f"BEA API Error: {res['Error']}")
    return res


def _extract_data_rows(data: dict) -> list[dict]:
    res = data["BEAAPI"]["Results"]
    if isinstance(res, list):
        for item in res:
            if isinstance(item, dict) and "Data" in item:
                return item["Data"]
        raise RuntimeError("Results 리스트 안에 Data 가 없음")
    if isinstance(res, dict) and "Data" in res:
        return res["Data"]
    raise RuntimeError("Data 를 찾지 못함")


# ==================================================================
# TableID / Year 탐색
# ==================================================================
def list_table_ids() -> pd.DataFrame:
    log.info("InputOutput 데이터셋의 TableID 목록을 조회합니다…")
    data = _get({
        "method": "GetParameterValues",
        "datasetname": DATASET,
        "ParameterName": "TableID",
    })
    res = _unwrap_results(data)
    return pd.DataFrame(res["ParamValue"])


def pick_table_id(df: pd.DataFrame, table_type: str, level: str) -> int | None:
    """원하는 (종류, 수준) 조합의 TableID 를 골라낸다. 없으면 None."""
    text_col = "Desc" if "Desc" in df.columns else df.columns[-1]
    desc_upper = df[text_col].astype(str).str.upper()

    rule = TYPE_RULES[table_type]
    # include 의 모든 토큰이 들어있어야 매칭 (AND) — IxI 와 IxC 같은 변형 구분에 필요
    mask_inc = desc_upper.apply(
        lambda s: all(tok in s for tok in rule["include"])
    )
    if rule["exclude"]:
        mask_exc = desc_upper.apply(
            lambda s: any(tok in s for tok in rule["exclude"])
        )
        mask_type = mask_inc & (~mask_exc)
    else:
        mask_type = mask_inc

    mask_level = desc_upper.apply(
        lambda s: any(tok in s for tok in LEVEL_TOKENS[level])
    )
    cand = df[mask_type & mask_level]
    if cand.empty:
        return None

    # "Before redefinitions" 우선
    pref = cand[cand[text_col].astype(str).str.contains("Before", case=False, na=False)]
    picked = (pref if not pref.empty else cand).iloc[0]
    tid = int(picked["Key"])
    log.info(f"[{table_type}/{level}] 선택된 TableID = {tid}  ({picked[text_col]})")
    return tid


def list_available_years(table_id: int) -> list[int]:
    data = _get({
        "method": "GetParameterValuesFiltered",
        "datasetname": DATASET,
        "TargetParameter": "Year",
        "TableID": str(table_id),
    })
    res = _unwrap_results(data)
    rows = res["ParamValue"]
    return sorted({int(r["Key"]) for r in rows})


# ==================================================================
# Year 단위 다운로드 + pivot
# ==================================================================
def fetch_year(table_id: int, year: int, raw_dir: Path) -> pd.DataFrame:
    raw_path = raw_dir / f"bea_{year}.json"
    if raw_path.exists() and raw_path.stat().st_size > 0:
        with raw_path.open(encoding="utf-8") as f:
            data = json.load(f)
    else:
        data = _get({
            "method": "GetData",
            "datasetname": DATASET,
            "TableID": str(table_id),
            "Year": str(year),
        })
        with raw_path.open("w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
        time.sleep(SLEEP_BETWEEN_CALLS)

    rows = _extract_data_rows(data)
    df = pd.DataFrame(rows)
    if df.empty:
        raise RuntimeError(f"{year}년 Data 가 비어 있음")
    return df


# BEA 부가축 코드 식별 정규식
# - 행에서 X 제외: V*(부가가치) / T0*(합계·세금)
# - 열에서 X 제외: F+digit (최종수요) / T0* (합계)
_BEA_AUX_ROW = re.compile(r"^(V\d|VA|T0)")
_BEA_AUX_COL = re.compile(r"^(F\d|T0)")


def extract_x_matrix(
    year: int,
    level: str = "summary",
    table: str = "use",
    drop_ghosts: bool = False,
) -> pd.DataFrame:
    """raw json 에서 한국은행 식 X(중간거래 행렬)만 잘라낸다.

    Parameters
    ----------
    year       : 연도 (raw 캐시에 있어야 함)
    level      : 'sector' / 'summary' (DETAIL 미발표)
    table      : 'use' / 'make' / 'tr_ixi' / 'tr_ixc' / 'tr_cxc'
    drop_ghosts: True 면 'Used'(scrap) 와 'Other'(noncomparable imports) 행도 제외.
                 False(기본)는 한국은행 통합중분류와 마찬가지로 보존.

    Returns
    -------
    DataFrame (MultiIndex (Code, Name) × MultiIndex (Code, Name))
        부가가치 행·최종수요 열을 모두 빼낸 직사각 X.
    """
    p = RAW_ROOT / f"{table}_{level}" / f"bea_{year}.json"
    with p.open(encoding="utf-8") as f:
        data = json.load(f)
    raw = pd.DataFrame(_extract_data_rows(data))
    raw["DataValue"] = pd.to_numeric(
        raw["DataValue"].astype(str).str.replace(",", "", regex=False),
        errors="coerce",
    )

    # universe(전체 행/열 코드)를 raw 에서 먼저 정의해야 GFGD/HS 처럼 데이터가
    # 부가축에만 있는 commodity 도 X 의 행으로 보존된다.
    all_rows = raw.drop_duplicates("RowCode")[["RowCode", "RowDescr"]]
    all_cols = raw.drop_duplicates("ColCode")[["ColCode", "ColDescr"]]
    keep_rows = all_rows[
        ~all_rows["RowCode"].astype(str).str.match(_BEA_AUX_ROW, na=False)
    ]
    keep_cols = all_cols[
        ~all_cols["ColCode"].astype(str).str.match(_BEA_AUX_COL, na=False)
    ]
    if drop_ghosts:
        keep_rows = keep_rows[~keep_rows["RowCode"].isin(["Used", "Other"])]

    df = raw[
        raw["RowCode"].isin(keep_rows["RowCode"])
        & raw["ColCode"].isin(keep_cols["ColCode"])
    ]
    rmap = keep_rows.set_index("RowCode")["RowDescr"]
    cmap = keep_cols.set_index("ColCode")["ColDescr"]
    X = df.pivot_table(
        index="RowCode", columns="ColCode", values="DataValue", aggfunc="sum"
    )
    X = X.reindex(index=rmap.index, columns=cmap.index)
    X.index = pd.MultiIndex.from_arrays(
        [X.index, X.index.map(rmap)], names=["Code", "Name"]
    )
    X.columns = pd.MultiIndex.from_arrays(
        [X.columns, X.columns.map(cmap)], names=["Code", "Name"]
    )
    return X


def long_to_matrix(df: pd.DataFrame) -> pd.DataFrame:
    col_map = {c.lower(): c for c in df.columns}
    rc = col_map.get("rowcode", "RowCode")
    rd = col_map.get("rowdescr", col_map.get("rowdescription", "RowDescr"))
    cc = col_map.get("colcode", "ColCode")
    cd = col_map.get("coldescr", col_map.get("coldescription", "ColDescr"))
    dv = col_map.get("datavalue", "DataValue")

    df = df.copy()
    df[dv] = (
        df[dv].astype(str)
        .str.replace(",", "", regex=False)
        .str.replace("---", "", regex=False)
    )
    df[dv] = pd.to_numeric(df[dv], errors="coerce")

    row_order = df.drop_duplicates(rc)[[rc, rd]].rename(
        columns={rc: "Code", rd: "Name"})
    col_order = df.drop_duplicates(cc)[[cc, cd]].rename(
        columns={cc: "Code", cd: "Name"})

    mat = df.pivot_table(index=rc, columns=cc, values=dv, aggfunc="sum")
    mat = mat.reindex(index=row_order["Code"], columns=col_order["Code"])
    mat.index = pd.MultiIndex.from_frame(row_order)
    mat.columns = pd.MultiIndex.from_frame(col_order)
    return mat


# ==================================================================
# Excel writer
# ==================================================================
def write_matrix_to_sheet(ws, mat: pd.DataFrame, year: int,
                          table_type: str, level: str) -> None:
    bold = Font(name="Arial", bold=True)
    normal = Font(name="Arial")
    header_fill = PatternFill("solid", start_color="D9E1F2")
    thin = Side(border_style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    num_fmt = "#,##0;(#,##0);-"

    ws.cell(1, 1).value = (
        f"BEA — {TYPE_HUMAN_NAME[table_type]} — {year} "
        f"({level} level, Millions of USD)"
    )
    ws.cell(1, 1).font = Font(name="Arial", bold=True, size=13)

    ws.cell(3, 1).value = "Row Code"
    ws.cell(3, 2).value = "Row Name"
    for c in (1, 2):
        ws.cell(3, c).font = bold
        ws.cell(3, c).fill = header_fill

    for j, (code, name) in enumerate(mat.columns.to_list(), start=3):
        ws.cell(2, j).value = code
        ws.cell(3, j).value = name
        for r in (2, 3):
            cell = ws.cell(r, j)
            cell.font = bold
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
            cell.border = border

    for i, ((rcode, rname), values) in enumerate(
        zip(mat.index.to_list(), mat.itertuples(index=False, name=None)),
        start=4,
    ):
        ws.cell(i, 1).value = rcode
        ws.cell(i, 2).value = rname
        ws.cell(i, 1).font = bold
        ws.cell(i, 2).font = normal
        ws.cell(i, 1).fill = header_fill
        ws.cell(i, 2).fill = header_fill
        for j, v in enumerate(values, start=3):
            cell = ws.cell(i, j)
            cell.value = None if pd.isna(v) else float(v)
            cell.number_format = num_fmt
            cell.font = normal
            cell.border = border

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 48
    for j in range(3, 3 + mat.shape[1]):
        ws.column_dimensions[get_column_letter(j)].width = 14
    ws.freeze_panes = "C4"
    ws.row_dimensions[3].height = 40


def build_workbook(year_to_matrix: dict[int, pd.DataFrame],
                   table_type: str, level: str, out_path: Path) -> None:
    wb = Workbook()
    summ = wb.active
    summ.title = "README"
    summ.cell(1, 1).value = (
        f"{TYPE_HUMAN_NAME[table_type]} — {level} level — "
        f"{min(year_to_matrix)}~{max(year_to_matrix)}"
    )
    summ.cell(1, 1).font = Font(name="Arial", bold=True, size=14)
    summ.cell(3, 1).value = (
        "각 시트는 한 해의 매트릭스 (Commodity × Industry + 최종수요 + 부가가치)입니다."
    )
    summ.cell(4, 1).value = "단위: Millions of USD (백만 달러)"
    summ.cell(5, 1).value = (
        "출처: U.S. Bureau of Economic Analysis (BEA), Input-Output Accounts, API."
    )
    summ.cell(6, 1).value = f"표 종류: {TYPE_HUMAN_NAME[table_type]}"
    summ.cell(7, 1).value = f"분류 수준: {level}"
    summ.cell(8, 1).value = "스크립트: bea_io_download.py (v4)"
    summ.cell(10, 1).value = "수록 연도"
    summ.cell(10, 1).font = Font(name="Arial", bold=True)
    for i, y in enumerate(sorted(year_to_matrix), start=11):
        summ.cell(i, 1).value = y
    summ.column_dimensions["A"].width = 90

    for year in sorted(year_to_matrix):
        ws = wb.create_sheet(str(year))
        write_matrix_to_sheet(ws, year_to_matrix[year], year, table_type, level)

    wb.save(out_path)
    log.info(f"  → 저장 완료: {out_path}")


# ==================================================================
# (종류, 수준) 단위 메인 루틴
# ==================================================================
def process_combo(table_type: str, level: str,
                  tbl_df: pd.DataFrame) -> None:
    tag = f"{table_type}/{level}"
    log.info(f"================ {tag} ================")
    table_id = pick_table_id(tbl_df, table_type, level)
    if table_id is None:
        log.warning(f"[{tag}] 매칭되는 TableID 없음. 건너뜀.")
        return

    avail = list_available_years(table_id)
    if not avail:
        log.warning(f"[{tag}] 가용 연도가 없습니다. 건너뜀.")
        return
    log.info(f"[{tag}] API 가용 연도: {avail[0]} ~ {avail[-1]} ({len(avail)}개)")

    end = END_YEAR or avail[-1]
    target_years = [y for y in avail if START_YEAR <= y <= end]
    if not target_years:
        log.warning(
            f"[{tag}] START_YEAR={START_YEAR}, END_YEAR={end} 범위에 해당하는 "
            f"가용 연도가 없습니다. (DETAIL 은 보통 벤치마크 연도만 제공: "
            f"{', '.join(map(str, avail))})"
        )
        return
    log.info(f"[{tag}] 받을 연도: {target_years}")

    raw_dir = RAW_ROOT / f"{table_type.lower()}_{level.lower()}"
    raw_dir.mkdir(exist_ok=True)

    year_to_matrix: dict[int, pd.DataFrame] = {}
    for y in target_years:
        try:
            log.info(f"  [{tag} {y}] 다운로드/파싱 중…")
            df = fetch_year(table_id, y, raw_dir)
            mat = long_to_matrix(df)
            year_to_matrix[y] = mat
            log.info(f"  [{tag} {y}] OK — shape={mat.shape}")
        except Exception as e:
            log.error(f"  [{tag} {y}] 실패: {e}")

    if not year_to_matrix:
        log.warning(f"[{tag}] 단 한 해도 성공하지 못했습니다. 다음으로 진행.")
        return

    out_path = OUTPUT_DIR / (
        f"{TYPE_FILE_PREFIX[table_type]}_all_years_{level.lower()}.xlsx"
    )
    build_workbook(year_to_matrix, table_type, level, out_path)


# ==================================================================
# 산업분류표 (Industry / Commodity Classification) 추출
# ==================================================================
# Use 형 표는 같은 LEVEL 에서는 USE/IMPORT/DOMESTIC 모두 동일한 코드 체계를 쓰므로
# raw json 들의 union 으로 분류 사전을 구성한다. 추가 API 호출이 필요 없다.
def _classify_code(code: str) -> str:
    """BEA Row/Col 코드를 카테고리로 분류한다.

    Industry/Commodity = 실제 산업·재화·서비스 (분석 대상)
    ValueAdded         = V001/V002/V003, VABAS, VAPRO
    Tax/Subsidy        = T00TOP / T00SUB / T00OTOP / T00OSUB
    Total              = T001 / T005 / T018 / T019 / 기타 T 시작 합계
    FinalDemand        = F + 숫자 시작 (F010, F02E, F100 ...). FIRE 처럼 F+문자 는 제외.
    Special (Scrap/Used)         = 'Used'
    Special (Noncomparable)      = 'Other'
    """
    c = str(code).strip()
    if c == "Used":
        return "Special (Scrap/Used)"
    if c == "Other":
        return "Special (Noncomparable)"
    if c in ("V001", "V002", "V003") or c.startswith("VA"):
        return "ValueAdded"
    if c.startswith("T00"):
        return "Tax/Subsidy"
    if c in ("T001", "T005", "T018", "T019"):
        return "Total"
    if c.startswith("T") and len(c) > 1 and c[1].isdigit():
        return "Total"
    if len(c) >= 2 and c[0] == "F" and c[1].isdigit():
        return "FinalDemand"
    return "Industry/Commodity"


def collect_codes_from_raw(raw_dir: Path) -> tuple[dict[str, str], dict[str, str]]:
    """raw_dir 안의 모든 bea_YYYY.json 을 합쳐 (코드→이름) 사전을 행/열별로 반환."""
    rows: dict[str, str] = {}
    cols: dict[str, str] = {}
    for jp in sorted(raw_dir.glob("bea_*.json")):
        try:
            with jp.open(encoding="utf-8") as f:
                data = json.load(f)
            for r in _extract_data_rows(data):
                rcode = str(r.get("RowCode", "")).strip()
                rdesc = str(r.get("RowDescr", "")).strip()
                ccode = str(r.get("ColCode", "")).strip()
                cdesc = str(r.get("ColDescr", "")).strip()
                if rcode:
                    rows.setdefault(rcode, rdesc)
                if ccode:
                    cols.setdefault(ccode, cdesc)
        except Exception as e:
            log.warning(f"  [분류] {jp.name} 스킵: {e}")
    return rows, cols


_TYPE_ORDER = {
    "Industry/Commodity": 0,
    "Special (Scrap/Used)": 1,
    "Special (Noncomparable)": 2,
    "ValueAdded": 3,
    "Tax/Subsidy": 4,
    "FinalDemand": 5,
    "Total": 6,
}


def _codes_to_df(d: dict[str, str]) -> pd.DataFrame:
    df = pd.DataFrame({"Code": list(d.keys()), "Name": list(d.values())})
    df["Type"] = df["Code"].map(_classify_code)
    df["_ord"] = df["Type"].map(_TYPE_ORDER).fillna(99).astype(int)
    df = df.sort_values(["_ord", "Code"]).drop(columns="_ord").reset_index(drop=True)
    return df


def parse_bea_naics_hierarchy(xlsx_path: Path) -> pd.DataFrame:
    """BEA DETAIL xlsx 의 'NAICS Codes' 시트를 파싱해 4계층 hierarchy 를 반환.

    BEA 가 발표하는 시트 구조 (들여쓰기 식):
        col 1 = Sector code (sector 행에서만 채워짐)
        col 2 = Summary code (summary 행에서만 채워짐) | Sector name (sector 행)
        col 3 = U.Summary code (u.summary 행에서만)    | Summary name (summary 행)
        col 4 = Detail code (detail 행에서만)          | U.Summary name (u.summary 행)
        col 5 = Detail name (detail 행)

    각 detail 행에 대해 부모 (sector / summary / u.summary) 를 currently-active
    상태에서 채워 한 행으로 정리한다.

    Returns DataFrame with columns:
        sector_code, sector_name, summary_code, summary_name,
        u_summary_code, u_summary_name, detail_code, detail_name
    """
    import openpyxl as _xl
    wb = _xl.load_workbook(str(xlsx_path), read_only=True, data_only=True)
    if "NAICS Codes" not in wb.sheetnames:
        raise ValueError(f"'NAICS Codes' 시트가 {xlsx_path.name} 에 없습니다.")
    ws = wb["NAICS Codes"]

    def _v(cell):
        """공백만 들어있는 셀도 None 으로 취급."""
        if cell.value is None:
            return None
        s = str(cell.value).strip()
        return s if s else None

    cur_sector = (None, None)
    cur_summary = (None, None)
    cur_u_summary = (None, None)
    out: list[dict] = []
    HEADER_TEXTS = {"Sector", "Summary", "U.Summary", "Detail",
                    "BEA Industry Code", "Industry Title"}

    for r in range(6, ws.max_row + 1):           # R1-R5 는 헤더, R6+ 데이터
        c1 = _v(ws.cell(r, 1))
        c2 = _v(ws.cell(r, 2))
        c3 = _v(ws.cell(r, 3))
        c4 = _v(ws.cell(r, 4))
        c5 = _v(ws.cell(r, 5))
        # 헤더 텍스트가 데이터로 끼어든 경우 무시
        if c1 in HEADER_TEXTS: c1 = None
        if c2 in HEADER_TEXTS: c2 = None
        if c3 in HEADER_TEXTS: c3 = None
        if c4 in HEADER_TEXTS: c4 = None

        if c1 is not None:
            cur_sector = (c1, c2 or "")
            cur_summary = (None, None)
            cur_u_summary = (None, None)
        elif c2 is not None:
            cur_summary = (c2, c3 or "")
            cur_u_summary = (None, None)
        elif c3 is not None:
            cur_u_summary = (c3, c4 or "")
        elif c4 is not None:
            out.append({
                "sector_code":     cur_sector[0],
                "sector_name":     cur_sector[1],
                "summary_code":    cur_summary[0],
                "summary_name":    cur_summary[1],
                "u_summary_code":  cur_u_summary[0],
                "u_summary_name":  cur_u_summary[1],
                "detail_code":     c4,
                "detail_name":     c5 or "",
            })
    return pd.DataFrame(out)


def _find_naics_source() -> Path | None:
    """OUTPUT_DIR 의 DETAIL 워크북 중 'NAICS Codes' 를 가진 첫 파일."""
    candidates = [
        "IOUse_Before_Redefinitions_PRO_DET.xlsx",
        "IOUse_After_Redefinitions_PRO_DET.xlsx",
        "IOMake_Before_Redefinitions_DET.xlsx",
        "IOMake_After_Redefinitions_DET.xlsx",
        "ImportMatrices_Before_Redefinitions_DET_2017.xlsx",
    ]
    for fn in candidates:
        p = OUTPUT_DIR / fn
        if p.exists():
            return p
    return None


def build_classification_workbook(out_path: Path) -> None:
    """BEA 산업분류표 워크북을 만든다.

    핵심 데이터: BEA DETAIL xlsx 의 'NAICS Codes' 시트 (= Sector → Summary →
    U.Summary → Detail 공식 hierarchy). DETAIL 워크북이 없으면 SUMMARY/SECTOR
    raw json 만으로 limited 분류표를 만든다.
    """
    bold = Font(name="Arial", bold=True)
    header_fill = PatternFill("solid", start_color="D9E1F2")

    wb = Workbook()
    readme = wb.active
    readme.title = "README"
    readme.cell(1, 1).value = "BEA Input-Output 산업분류표 (Sector → Summary → U.Summary → Detail 계층)"
    readme.cell(1, 1).font = Font(name="Arial", bold=True, size=14)
    readme_lines = [
        "",
        "BEA 가 발표하는 4 계층 산업 hierarchy:",
        "  · Sector       (~16개)   가장 거친 분류. 코드 예: 11, 21, 31G, 44RT.",
        "  · Summary      (~71개)   중간 분류. 코드 예: 111CA, 113FF, 211, 311FT.",
        "  · U.Summary    (~140개)  Summary 와 Detail 사이의 사용자 그룹. 코드 예: 111, 112, 211.",
        "  · Detail       (~405개)  가장 세분. 코드 예: 1111A0, 111200, 211000.",
        "",
        "시트 구성:",
        "  · Hierarchy    — 모든 Detail 행 한 줄씩 + 그 부모 (Sector/Summary/U.Summary). lookup 용",
        "  · Sector       — Sector 코드만 (이름 포함)",
        "  · Summary      — Summary 코드 + 부모 Sector",
        "  · U.Summary    — U.Summary 코드 + 부모 Summary + 부모 Sector",
        "  · Detail       — Detail 코드 + 모든 부모",
        "  · BEA_부가축   — Use/Make 표에서 본 V*/T*/F* 등 비-산업 코드 (분석 시 제외 대상)",
        "",
        "출처: BEA DETAIL xlsx ('NAICS Codes' 시트). DETAIL 데이터 없으면 SUMMARY/SECTOR raw json fallback.",
        "스크립트: bea_io_download.py — build_classification_workbook() / parse_bea_naics_hierarchy()",
    ]
    for i, line in enumerate(readme_lines, start=2):
        readme.cell(i, 1).value = line
    readme.column_dimensions["A"].width = 100

    naics_path = _find_naics_source()
    if naics_path is None:
        log.warning(
            "BEA DETAIL xlsx 가 없어 hierarchy 시트를 만들 수 없습니다. "
            "`download_bea_static_files()` 를 먼저 실행하세요. "
            "raw json 만으로 limited 분류표를 작성합니다."
        )
        # Fallback: 기존 raw-json 만의 분류 (이전 버전과 동일한 구조)
        return _build_classification_legacy(out_path, readme, readme_lines, wb)

    log.info(f"[분류] hierarchy 소스: {naics_path.name}")
    h = parse_bea_naics_hierarchy(naics_path)
    log.info(f"[분류] hierarchy: detail {len(h)}개 행")

    # Helper to write a level sheet with header row.
    def _write_level_sheet(name: str, df: pd.DataFrame, columns: list[str]) -> None:
        ws = wb.create_sheet(name)
        ws.cell(1, 1).value = f"BEA {name} — 총 {len(df)}개"
        ws.cell(1, 1).font = Font(name="Arial", bold=True, size=12)
        for j, h_ in enumerate(columns, start=1):
            c = ws.cell(3, j); c.value = h_; c.font = bold; c.fill = header_fill
        for i, row in df.reset_index(drop=True).iterrows():
            for j, col in enumerate(columns, start=1):
                ws.cell(4 + i, j).value = row.get(col)
        widths = {"sector_code": 10, "summary_code": 12, "u_summary_code": 14,
                  "detail_code": 12,
                  "sector_name": 50, "summary_name": 50,
                  "u_summary_name": 50, "detail_name": 60}
        for j, col in enumerate(columns, start=1):
            from openpyxl.utils import get_column_letter as _gcl
            ws.column_dimensions[_gcl(j)].width = widths.get(col, 18)
        ws.freeze_panes = "A4"

    # 1) Hierarchy (every detail row + parents) — main lookup sheet
    _write_level_sheet(
        "Hierarchy", h,
        ["sector_code", "sector_name", "summary_code", "summary_name",
         "u_summary_code", "u_summary_name", "detail_code", "detail_name"],
    )

    # 2) Sector — unique sector codes
    sec = h.drop_duplicates("sector_code")[["sector_code", "sector_name"]]
    _write_level_sheet("Sector", sec, ["sector_code", "sector_name"])

    # 3) Summary — unique summary codes + parent sector
    summ = h.drop_duplicates("summary_code")[
        ["summary_code", "summary_name", "sector_code", "sector_name"]
    ]
    _write_level_sheet("Summary", summ,
                       ["summary_code", "summary_name", "sector_code", "sector_name"])

    # 4) U.Summary
    usum = h.drop_duplicates("u_summary_code")[
        ["u_summary_code", "u_summary_name",
         "summary_code", "summary_name", "sector_code", "sector_name"]
    ]
    _write_level_sheet("U.Summary", usum,
                       ["u_summary_code", "u_summary_name",
                        "summary_code", "summary_name", "sector_code", "sector_name"])

    # 5) Detail
    _write_level_sheet("Detail", h,
                       ["detail_code", "detail_name",
                        "u_summary_code", "u_summary_name",
                        "summary_code", "summary_name",
                        "sector_code", "sector_name"])

    # 6) BEA_부가축 — non-industry codes from raw json (V*/T*/F*/Special)
    aux_rows: dict[str, str] = {}
    aux_cols: dict[str, str] = {}
    for level in LEVELS:
        for ttype in TABLE_TYPES:
            raw_dir = RAW_ROOT / f"{ttype.lower()}_{level.lower()}"
            if not raw_dir.exists():
                continue
            r, c = collect_codes_from_raw(raw_dir)
            aux_rows.update(r)
            aux_cols.update(c)
    aux_row_list = [(k, v, _classify_code(k))
                    for k, v in sorted(aux_rows.items())
                    if _classify_code(k) != "Industry/Commodity"]
    aux_col_list = [(k, v, _classify_code(k))
                    for k, v in sorted(aux_cols.items())
                    if _classify_code(k) != "Industry/Commodity"]
    if aux_row_list or aux_col_list:
        ws = wb.create_sheet("BEA_부가축")
        ws.cell(1, 1).value = (
            f"BEA Use/Make 표의 부가축 코드 (산업·재화 아닌 행/열) — "
            f"분석 시 X 매트릭스에서 제외 대상"
        )
        ws.cell(1, 1).font = Font(name="Arial", bold=True, size=12)
        ws.cell(3, 1).value = "[행 (부가가치/세금/합계)]"; ws.cell(3, 1).font = bold
        ws.cell(3, 5).value = "[열 (최종수요/합계)]";    ws.cell(3, 5).font = bold
        for j, h_ in enumerate(["Code", "Name", "Type"], start=1):
            c = ws.cell(4, j); c.value = h_; c.font = bold; c.fill = header_fill
        for j, h_ in enumerate(["Code", "Name", "Type"], start=5):
            c = ws.cell(4, j); c.value = h_; c.font = bold; c.fill = header_fill
        for i, (code, name, t) in enumerate(aux_row_list):
            ws.cell(5 + i, 1).value = code
            ws.cell(5 + i, 2).value = name
            ws.cell(5 + i, 3).value = t
        for i, (code, name, t) in enumerate(aux_col_list):
            ws.cell(5 + i, 5).value = code
            ws.cell(5 + i, 6).value = name
            ws.cell(5 + i, 7).value = t
        ws.column_dimensions["A"].width = 12
        ws.column_dimensions["B"].width = 50
        ws.column_dimensions["C"].width = 22
        ws.column_dimensions["D"].width = 2
        ws.column_dimensions["E"].width = 12
        ws.column_dimensions["F"].width = 50
        ws.column_dimensions["G"].width = 22
        ws.freeze_panes = "A5"

    # README 끝에 카운트 요약
    summary_row = len(readme_lines) + 3
    readme.cell(summary_row, 1).value = (
        f"수록: Sector {len(sec)} / Summary {len(summ)} / "
        f"U.Summary {len(usum)} / Detail {len(h)} 산업"
    )
    readme.cell(summary_row, 1).font = bold

    wb.save(out_path)
    log.info(f"  → 저장 완료: {out_path}")
    log.info(
        f"[분류] hierarchy 워크북: Sector {len(sec)} / Summary {len(summ)} / "
        f"U.Summary {len(usum)} / Detail {len(h)}"
    )


def _build_classification_legacy(
    out_path: Path,
    readme,
    readme_lines: list[str],
    wb: "Workbook",
) -> None:
    """DETAIL xlsx 가 없을 때 raw json 만으로 limited 분류표 만드는 폴백."""
    bold = Font(name="Arial", bold=True)
    header_fill = PatternFill("solid", start_color="D9E1F2")

    levels_written: list[tuple[str, int, int, int, int]] = []
    for level in LEVELS:
        rows_all: dict[str, str] = {}
        cols_all: dict[str, str] = {}
        for ttype in TABLE_TYPES:
            raw_dir = RAW_ROOT / f"{ttype.lower()}_{level.lower()}"
            if not raw_dir.exists():
                continue
            r, c = collect_codes_from_raw(raw_dir)
            for k, v in r.items():
                rows_all.setdefault(k, v)
            for k, v in c.items():
                cols_all.setdefault(k, v)
        if not rows_all and not cols_all:
            continue
        rdf = _codes_to_df(rows_all)
        cdf = _codes_to_df(cols_all)
        ws = wb.create_sheet(f"{level.title()}_legacy")
        ws.cell(1, 1).value = f"BEA — {level} (raw json 폴백) — 행 {len(rdf)} / 열 {len(cdf)}"
        ws.cell(1, 1).font = Font(name="Arial", bold=True, size=12)
        for j, h in enumerate(["Code", "Name", "Type"], start=1):
            c = ws.cell(3, j); c.value = h; c.font = bold; c.fill = header_fill
        for j, h in enumerate(["Code", "Name", "Type"], start=5):
            c = ws.cell(3, j); c.value = h; c.font = bold; c.fill = header_fill
        for i, r in rdf.iterrows():
            ws.cell(4 + i, 1).value = r["Code"]
            ws.cell(4 + i, 2).value = r["Name"]
            ws.cell(4 + i, 3).value = r["Type"]
        for i, r in cdf.iterrows():
            ws.cell(4 + i, 5).value = r["Code"]
            ws.cell(4 + i, 6).value = r["Name"]
            ws.cell(4 + i, 7).value = r["Type"]
        levels_written.append((level, len(rdf), len(cdf), 0, 0))

    summary_row = len(readme_lines) + 3
    readme.cell(summary_row, 1).value = "DETAIL xlsx 미존재 → legacy 시트만 작성 (hierarchy 없음)"
    readme.cell(summary_row, 1).font = bold

    wb.save(out_path)
    log.info(f"  → 저장 완료(legacy): {out_path}")


# ==================================================================
# BEA 정적 Excel 다운로드 (API 가 노출하지 않는 표용)
# ==================================================================
# 한국은행 Aᵐ 에 해당하는 "Use of Imported Commodities by Industries" 는
# BEA InputOutput API 가 노출하지 않는다. 대신 BEA industry-data 아카이브에
# 연도 묶음 단일 워크북(.xlsx)으로 발표된다. URL 패턴은
#   {BEA_STATIC_BASE}{stem}_{base_year}-{end_year}.xlsx
# 이고, 파일이 없는 URL 은 200 OK 의 HTML soft-404 를 돌려주므로
# Content-Type 에 'spreadsheet' 가 포함된 경우만 진짜 파일로 본다.
BEA_STATIC_BASE = "https://apps.bea.gov/industry/xls/io-annual/"
# (stem, base_year) for files that follow the {stem}_{base_year}-YYYY.xlsx pattern
# at SUMMARY level (year-range workbooks).
BEA_IMPORT_STEMS = [
    "ImportMatrices_Before_Redefinitions_SUM",
    "ImportMatrices_After_Redefinitions_SUM",
]
# DETAIL-level files have a fixed filename (no year-range suffix). Each is a
# multi-sheet workbook keyed by benchmark year (1997 / 2002 / 2007 / 2012 / 2017).
BEA_DETAIL_FILES = [
    "IOUse_Before_Redefinitions_PRO_DET.xlsx",
    "IOUse_After_Redefinitions_PRO_DET.xlsx",
    "IOMake_Before_Redefinitions_DET.xlsx",
    "IOMake_After_Redefinitions_DET.xlsx",
    "ImportMatrices_Before_Redefinitions_DET_2017.xlsx",
    "ImportMatrices_After_Redefinitions_DET_2017.xlsx",
]


def _bea_local_name(url: str) -> str:
    """BEA URL 의 파일명을 'bea_' + snake_case (약어는 대문자 유지) 로 변환.

    예) ImportMatrices_Before_Redefinitions_SUM_1997-2023.xlsx
        →  bea_import_matrices_before_redefinitions_SUM_1997-2023.xlsx
    """
    remote = url.rsplit("/", 1)[-1]
    parts = remote.split("_")
    out: list[str] = []
    for p in parts:
        if p.isupper() and len(p) <= 4:        # SUM/DET/SEC 같은 약어는 유지
            out.append(p)
        else:
            out.append(re.sub(r"([a-z])([A-Z])", r"\1_\2", p).lower())
    return "bea_" + "_".join(out)


def _bea_static_is_real(url: str) -> bool:
    try:
        req = urllib.request.Request(
            url, method="HEAD", headers={"User-Agent": "Mozilla/5.0"}
        )
        with urllib.request.urlopen(req, timeout=15) as r:
            return "spreadsheet" in r.headers.get("Content-Type", "").lower()
    except Exception:
        return False


def find_latest_bea_static(stem: str, base_year: int = 1997) -> str | None:
    """{stem}_{base_year}-YYYY.xlsx 의 최신 YYYY 를 HEAD 로 탐색."""
    import datetime
    this_year = datetime.date.today().year
    for y in range(this_year + 1, base_year, -1):
        url = f"{BEA_STATIC_BASE}{stem}_{base_year}-{y}.xlsx"
        if _bea_static_is_real(url):
            return url
    return None


def _download_one(url: str, out: Path, *, force: bool = False) -> Path | None:
    """단일 BEA 정적 파일 다운로드 (이미 있으면 건너뜀)."""
    if out.exists() and out.stat().st_size > 100_000 and not force:
        log.info(f"[정적] {out.name} 이미 존재 — 건너뜀 ({out.stat().st_size:,} bytes)")
        return out
    log.info(f"[정적] {url} → {out.name}")
    try:
        req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
        with urllib.request.urlopen(req, timeout=REQUEST_TIMEOUT) as resp:
            content = resp.read()
            ctype = resp.headers.get("Content-Type", "").lower()
        if "spreadsheet" not in ctype:
            log.warning(f"  [정적] content-type={ctype} — soft-404 가능. 건너뜀.")
            return None
        out.write_bytes(content)
        log.info(f"  [정적] 저장 완료: {out.stat().st_size:,} bytes")
        return out
    except Exception as e:
        log.error(f"  [정적] 다운로드 실패: {e}")
        return None


def download_bea_static_files(force: bool = False) -> list[Path]:
    """BEA 정적 워크북 (SUMMARY Import 다년도 + DETAIL Use/Make/Import) 다운로드.
    이미 존재하면 건너뜀."""
    saved: list[Path] = []
    # SUMMARY-level Import: latest year-range file ({stem}_1997-YYYY.xlsx)
    for stem in BEA_IMPORT_STEMS:
        url = find_latest_bea_static(stem)
        if url is None:
            log.warning(f"[정적] {stem} 의 최신 파일을 찾지 못했습니다.")
            continue
        out = OUTPUT_DIR / _bea_local_name(url)
        result = _download_one(url, out, force=force)
        if result is not None:
            saved.append(result)
    # DETAIL-level files: BEA filename preserved as-is (no _bea_local_name).
    for fn in BEA_DETAIL_FILES:
        url = BEA_STATIC_BASE + fn
        out = OUTPUT_DIR / fn
        result = _download_one(url, out, force=force)
        if result is not None:
            saved.append(result)
    return saved


# ==================================================================
# main
# ==================================================================
def main() -> None:
    log.info("=== BEA Input-Output 다운로드 시작 (v4 multi-type × multi-level) ===")
    tbl_df = list_table_ids()
    log.info(f"InputOutput 데이터셋의 TableID {len(tbl_df)} 개를 받아왔습니다.")

    log.info(f"받을 조합: {[(t, l) for t in TABLE_TYPES for l in LEVELS]}")

    for t in TABLE_TYPES:
        for l in LEVELS:
            try:
                process_combo(t, l, tbl_df)
            except Exception as e:
                log.error(f"[{t}/{l}] 치명적 오류로 건너뜀: {e}")

    log.info("=== Import Matrix 정적 Excel 다운로드 (API 미제공) ===")
    try:
        download_bea_static_files()
    except Exception as e:
        log.error(f"정적 Excel 다운로드 실패: {e}")

    classify_path = OUTPUT_DIR / "bea_industry_classification.xlsx"
    try:
        log.info("=== 산업분류표 추출 ===")
        build_classification_workbook(classify_path)
    except Exception as e:
        log.error(f"분류표 생성 실패: {e}")

    log.info("=== 전체 완료 ===")


if __name__ == "__main__":
    main()
